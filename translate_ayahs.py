import openpyxl
import os
import requests

# --- CONFIGURATION ---
EXCEL_FILE = 'vocab.xlsx'
SHEET_NAME = "Vocab with Examples"

# Inputs
TXT_SOURCE_ZILIO = 'quran_zilio_formatted_with_checks.txt'
API_URL_PICCARDO = "http://api.alquran.cloud/v1/quran/it.piccardo"
API_URL_SAHIH = "http://api.alquran.cloud/v1/quran/en.sahih"

# Column Headers
REF_HEADER = "Ayahref"
HEADER_ZILIO = "Meaning & Translation in Italian (Ida-Zilio)"
HEADER_PICCARDO = "Meaning & Translation in Italian (Hamza Roberto Piccardo)"
HEADER_SAHIH = "Meaning & Translation (Sahih International)"

def automate_quran_translation():
    
    # --- STEP 1: LOAD ZILIO (FROM TXT) ---
    print(f"1. Loading Zilio translation from '{TXT_SOURCE_ZILIO}'...")
    zilio_map = {}
    
    if os.path.exists(TXT_SOURCE_ZILIO):
        with open(TXT_SOURCE_ZILIO, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line: continue
                # Format: "2:14 Text..."
                parts = line.split(" ", 1)
                if len(parts) == 2:
                    zilio_map[parts[0].strip()] = parts[1].strip()
                elif len(parts) == 1:
                    zilio_map[parts[0].strip()] = "[TEXT MISSING]"
        print(f"   - Loaded {len(zilio_map)} verses for Zilio.")
    else:
        print(f"   - WARNING: File '{TXT_SOURCE_ZILIO}' not found. Zilio column will be empty.")

    # --- STEP 2: LOAD PICCARDO (FROM API) ---
    print("2. Loading Piccardo translation from API...")
    piccardo_map = {}
    try:
        response = requests.get(API_URL_PICCARDO)
        response.raise_for_status()
        data = response.json()['data']['surahs']
        for surah in data:
            surah_num = str(surah['number'])
            for ayah in surah['ayahs']:
                key = f"{surah_num}:{ayah['numberInSurah']}"
                piccardo_map[key] = ayah['text']
        print(f"   - Loaded {len(piccardo_map)} verses for Piccardo.")
    except Exception as e:
        print(f"   - Error fetching Piccardo API: {e}")

    # --- STEP 3: LOAD SAHIH INTERNATIONAL (FROM API) ---
    print("3. Loading Sahih International translation from API...")
    sahih_map = {}
    try:
        response = requests.get(API_URL_SAHIH)
        response.raise_for_status()
        data = response.json()['data']['surahs']
        for surah in data:
            surah_num = str(surah['number'])
            for ayah in surah['ayahs']:
                key = f"{surah_num}:{ayah['numberInSurah']}"
                sahih_map[key] = ayah['text']
        print(f"   - Loaded {len(sahih_map)} verses for Sahih International.")
    except Exception as e:
        print(f"   - Error fetching Sahih API: {e}")

    # --- STEP 4: OPEN EXCEL ---
    print(f"4. Opening Excel file '{EXCEL_FILE}'...")
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        print(f"Error: Excel file '{EXCEL_FILE}' not found.")
        return

    if SHEET_NAME not in wb.sheetnames:
        print(f"Error: Sheet '{SHEET_NAME}' not found.")
        return
    
    ws = wb[SHEET_NAME]

    # --- STEP 5: SETUP COLUMNS ---
    print("5. Setting up columns...")
    
    cols = {
        'ref': None,
        'zilio': None,
        'piccardo': None,
        'sahih': None
    }
    
    # Scan existing headers
    for cell in ws[1]:
        val = str(cell.value).strip() if cell.value else ""
        if val == REF_HEADER:
            cols['ref'] = cell.column
        elif val == HEADER_ZILIO:
            cols['zilio'] = cell.column
        elif val == HEADER_PICCARDO:
            cols['piccardo'] = cell.column
        elif val == HEADER_SAHIH:
            cols['sahih'] = cell.column

    if cols['ref'] is None:
        print(f"Error: Column '{REF_HEADER}' not found.")
        return

    # Create missing columns dynamically
    # We use a helper function to create columns if they don't exist
    def create_col_if_missing(col_key, header_name):
        if cols[col_key] is None:
            cols[col_key] = ws.max_column + 1
            ws.cell(row=1, column=cols[col_key]).value = header_name
            print(f"   - Created column: {header_name}")

    create_col_if_missing('zilio', HEADER_ZILIO)
    create_col_if_missing('piccardo', HEADER_PICCARDO)
    create_col_if_missing('sahih', HEADER_SAHIH)

    # --- STEP 6: FILL DATA ---
    print("6. Filling rows...")
    stats = {'zilio': 0, 'piccardo': 0, 'sahih': 0}

    for row in range(2, ws.max_row + 1):
        ref_cell = ws.cell(row=row, column=cols['ref'])
        
        raw_ref = ref_cell.value
        if raw_ref:
            # Clean Reference: " 2 : 102 " -> "2:102"
            clean_ref = str(raw_ref).replace(" ", "").replace(":", ":").strip()
            
            # Write Zilio
            if zilio_map.get(clean_ref):
                ws.cell(row=row, column=cols['zilio']).value = zilio_map[clean_ref]
                stats['zilio'] += 1
            
            # Write Piccardo
            if piccardo_map.get(clean_ref):
                ws.cell(row=row, column=cols['piccardo']).value = piccardo_map[clean_ref]
                stats['piccardo'] += 1

            # Write Sahih
            if sahih_map.get(clean_ref):
                ws.cell(row=row, column=cols['sahih']).value = sahih_map[clean_ref]
                stats['sahih'] += 1

    # --- STEP 7: SAVE ---
    print(f"7. Saving to '{EXCEL_FILE}'...")
    try:
        wb.save(EXCEL_FILE)
        print(f"Done! Updated Zilio: {stats['zilio']}, Piccardo: {stats['piccardo']}, Sahih: {stats['sahih']}.")
    except PermissionError:
        print("ERROR: Please close the Excel file before running this script!")

if __name__ == "__main__":
    automate_quran_translation()