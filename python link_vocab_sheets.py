import openpyxl
from openpyxl.utils import get_column_letter

# --- CONFIGURATION ---
FILE_NAME = 'database_with_examples.xlsx'
SHEET_SOURCE = 'Vocabulary Lists'      # Master list
SHEET_TARGET = 'Vocab with Examples'   # Sheet to fill

# Target Columns
COL_AYAHREF = "Ayahref"
COL_WORD_ID = "WordID"
COL_WORD = "Word"
COL_TARGET_MEANING = "Meaning & Translation in Italian (Hamza Roberto Piccardo)"

# Source Columns
COL_SOURCE_ID = "Id"
COL_SOURCE_WORD = "Word"
COL_SOURCE_MEANING = "Meaning (Italian)"

def normalize_text(text):
    if text is None: return ""
    return str(text).strip()

def link_sheets_ordered():
    print(f"Opening '{FILE_NAME}'...")
    try:
        wb = openpyxl.load_workbook(FILE_NAME)
    except FileNotFoundError:
        print(f"Error: File '{FILE_NAME}' not found.")
        return

    if SHEET_SOURCE not in wb.sheetnames or SHEET_TARGET not in wb.sheetnames:
        print("Error: Missing required sheets.")
        return

    ws_src = wb[SHEET_SOURCE]
    ws_tgt = wb[SHEET_TARGET]

    # --- STEP 1: MAP COLUMNS ---
    print("Mapping columns...")
    tgt_map = {str(c.value).strip(): c.column for c in ws_tgt[1] if c.value}
    
    # For source, we need letters (for formula) and indices (for reading)
    src_map = {}
    for c in ws_src[1]:
        if c.value:
            header = str(c.value).strip()
            src_map[header] = c.column_letter
            src_map[header + "_idx"] = c.column

    # Verify columns
    for c in [COL_AYAHREF, COL_WORD_ID, COL_WORD, COL_TARGET_MEANING]:
        if c not in tgt_map: return print(f"Error: Target col '{c}' missing.")
    for c in [COL_SOURCE_ID, COL_SOURCE_WORD, COL_SOURCE_MEANING]:
        if c not in src_map: return print(f"Error: Source col '{c}' missing.")

    # --- STEP 2: BUILD ID QUEUE ---
    # Structure: { 'WORD': [ID_1, ID_2, ID_3] }
    print("Building ID Queue...")
    word_queues = {}
    
    src_id_idx = src_map[COL_SOURCE_ID + "_idx"]
    src_word_idx = src_map[COL_SOURCE_WORD + "_idx"]

    for row in range(2, ws_src.max_row + 1):
        id_val = ws_src.cell(row=row, column=src_id_idx).value
        word_val = ws_src.cell(row=row, column=src_word_idx).value
        
        if id_val and word_val:
            clean_word = normalize_text(word_val)
            if clean_word not in word_queues:
                word_queues[clean_word] = []
            word_queues[clean_word].append(id_val)

    # --- STEP 3: PREPARE FORMULA ---
    src_id_col_idx = src_map[COL_SOURCE_ID + "_idx"]
    src_mean_col_idx = src_map[COL_SOURCE_MEANING + "_idx"]
    vlookup_idx = src_mean_col_idx - src_id_col_idx + 1

    # --- STEP 4: PROCESS TARGET ---
    print("Linking rows...")
    updates = 0
    # We use a pointer map to track which index we are at for each word
    # { 'WORD': 0 } -> means next time we see WORD, take index 0
    word_pointers = {} 
    
    previous_word_clean = None # For consecutive check

    for row in range(2, ws_tgt.max_row + 1):
        ayah_cell = ws_tgt.cell(row=row, column=tgt_map[COL_AYAHREF])
        word_cell = ws_tgt.cell(row=row, column=tgt_map[COL_WORD])
        
        ayah_val = normalize_text(ayah_cell.value)
        word_val_clean = normalize_text(word_cell.value)
        
        # Is this a Word Row?
        is_word_row = (ayah_val == "" and word_val_clean != "")

        if is_word_row:
            # CHECK CONSECUTIVE: If exact same word as previous row, skip logic
            if word_val_clean == previous_word_clean:
                continue

            # Retrieve ID List
            id_list = word_queues.get(word_val_clean)
            
            if id_list:
                # Get current pointer for this word (default 0)
                ptr = word_pointers.get(word_val_clean, 0)
                
                # If we still have IDs left for this word
                if ptr < len(id_list):
                    target_id = id_list[ptr]
                    
                    # Write ID
                    ws_tgt.cell(row=row, column=tgt_map[COL_WORD_ID]).value = target_id
                    
                    # Write Formula
                    id_col_letter = get_column_letter(tgt_map[COL_WORD_ID])
                    rng = f"'{SHEET_SOURCE}'!{src_map[COL_SOURCE_ID]}:{src_map[COL_SOURCE_MEANING]}"
                    formula = f"=VLOOKUP({id_col_letter}{row}, {rng}, {vlookup_idx}, FALSE)"
                    
                    ws_tgt.cell(row=row, column=tgt_map[COL_TARGET_MEANING]).value = formula
                    
                    # Increment pointer so next time we take the NEXT ID
                    word_pointers[word_val_clean] = ptr + 1
                    updates += 1
                else:
                    # We ran out of IDs! (e.g., Word appears 5 times in Examples but only 4 in List)
                    # We reuse the LAST ID to be safe
                    target_id = id_list[-1]
                    ws_tgt.cell(row=row, column=tgt_map[COL_WORD_ID]).value = target_id
                    # (Formula writing logic repeated...)
                    id_col_letter = get_column_letter(tgt_map[COL_WORD_ID])
                    rng = f"'{SHEET_SOURCE}'!{src_map[COL_SOURCE_ID]}:{src_map[COL_SOURCE_MEANING]}"
                    formula = f"=VLOOKUP({id_col_letter}{row}, {rng}, {vlookup_idx}, FALSE)"
                    ws_tgt.cell(row=row, column=tgt_map[COL_TARGET_MEANING]).value = formula
                    print(f"Warning: Word '{word_val_clean}' appears more times in Examples than in List. Reused last ID.")

            previous_word_clean = word_val_clean
        else:
            previous_word_clean = None

    print(f"Saving to '{FILE_NAME}'...")
    try:
        wb.save(FILE_NAME)
        print(f"Done! Linked {updates} words.")
    except PermissionError:
        print("ERROR: Close the Excel file first!")

if __name__ == "__main__":
    link_sheets_ordered()